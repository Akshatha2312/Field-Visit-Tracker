from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def build_excel_response(context):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_fill = PatternFill(fill_type="solid", fgColor="DCEBFA")
    bold_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="1F4E79")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    sheet.append(["Field Visit Tracker Report"])
    sheet.cell(row=1, column=1).font = title_font
    sheet.merge_cells("A1:E1")
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")

    sheet.append([])
    sheet.append(["Generated Date", context.get("generated_at", "")])
    sheet.append(["Applied Filters", ""])
    sheet.append(["From Date", context.get("from_date") or "All"])
    sheet.append(["To Date", context.get("to_date") or "All"])
    sheet.append(["Employee", context.get("employee_name") or "All"])
    sheet.append(["Visit Status", context.get("visit_status") or "All"])
    sheet.append([])
    sheet.append(["Summary", ""])
    sheet.append(["Total Attendance Records", context.get("total_attendance_records", 0)])
    sheet.append(["Total Present", context.get("total_present", 0)])
    sheet.append(["Total Client Visits", context.get("total_client_visits", 0)])
    sheet.append(["Pending Visits", context.get("pending_visits", 0)])
    sheet.append(["Completed Visits", context.get("completed_visits", 0)])
    sheet.append([])

    visits = context.get("visits", [])
    if visits:
        sheet.append(["Visit Details", ""])
        sheet.append(["Employee", "Client Name", "Company Name", "Visit Date", "Status"])
        for visit in visits:
            sheet.append(
                [
                    visit.employee.name if visit.employee else "-",
                    visit.client_name,
                    visit.company_name,
                    str(visit.visit_date),
                    visit.status,
                ]
            )
    else:
        sheet.append(["Visit Details", ""])
        sheet.append(["No records found."])

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border

    for row_idx in range(1, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.fill = title_fill
                cell.font = Font(bold=True, size=16, color="FFFFFF")
            elif row_idx in {3, 9, 15}:
                cell.font = Font(bold=True)
            elif row_idx == 11 and sheet.max_column >= 2:
                cell.font = Font(bold=True)

    header_rows = [2, 10, 16]
    for row_idx in header_rows:
        if row_idx <= sheet.max_row:
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="000000")

    for col_idx in range(1, sheet.max_column + 1):
        column = get_column_letter(col_idx)
        max_length = 0
        for row in sheet[column]:
            if row.value is not None:
                max_length = max(max_length, len(str(row.value)))
        adjusted_width = min(max_length + 2, 40)
        sheet.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="field_visit_tracker_report.xlsx"'
    return response
